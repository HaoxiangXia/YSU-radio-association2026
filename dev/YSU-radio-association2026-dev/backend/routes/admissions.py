import io
import json
import logging
import os
import shutil
import tempfile
import threading
import uuid
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, ConfigDict, Field, ValidationError

from config.recruitment import RecruitmentConfig, get_recruitment_config
from routes.recruitment_officers import (
    RecruitmentOfficerInfo,
    get_current_recruitment_officer,
)
from utils.security import admission_query_limiter, get_client_ip


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_ADMISSIONS_PATH = (
    REPOSITORY_ROOT.parent
    / "YSU-radio-association-private"
    / "admission-results.json"
)

router = APIRouter()
logger = logging.getLogger(__name__)

MAX_WORKBOOK_BYTES = 2 * 1024 * 1024
MAX_WORKBOOK_EXPANDED_BYTES = 20 * 1024 * 1024
MAX_WORKBOOK_ENTRIES = 300
MAX_ADMISSION_ROWS = 2000
PREVIEW_TTL = timedelta(minutes=30)
WORKSHEET_NAME = "录取名单"
WORKBOOK_COLUMNS = ("姓名", "学号", "申请手机号", "录取部门", "录取状态")


class AdmissionDataError(RuntimeError):
    pass


class AdmissionRecord(BaseModel):
    model_config = ConfigDict(extra="forbid", str_strip_whitespace=True)

    student_id: str = Field(alias="studentId", pattern=r"^\d{12}$")
    name: str = Field(min_length=1, max_length=30)
    phone: str = Field(pattern=r"^1[3-9]\d{9}$")
    department: str = Field(default="", max_length=100)
    status: Literal["已录取", "未录取"]


class AdmissionQuery(BaseModel):
    model_config = ConfigDict(extra="forbid", populate_by_name=True, str_strip_whitespace=True)

    student_id: str = Field(alias="studentId", pattern=r"^\d{12}$")
    phone: str = Field(pattern=r"^1[3-9]\d{9}$")


class AdmissionQueryResponse(BaseModel):
    name: str
    department: str
    status: Literal["已录取", "未录取"]


class PublishAdmissionsRequest(BaseModel):
    model_config = ConfigDict(extra="forbid", str_strip_whitespace=True)

    preview_id: str = Field(alias="previewId", min_length=32, max_length=64)


_admissions_by_student_id: dict[str, AdmissionRecord] | None = None
_admissions_path: Path | None = None
_admissions_lock = threading.RLock()
_previews: dict[str, dict] = {}
_preview_lock = threading.Lock()


def resolve_admissions_path(path: str | Path | None = None) -> Path:
    configured_path = path or os.environ.get("ADMISSIONS_DATA_PATH")
    if configured_path:
        resolved = Path(configured_path).expanduser()
        if not resolved.is_absolute():
            resolved = REPOSITORY_ROOT / resolved
        return resolved.resolve()
    return DEFAULT_ADMISSIONS_PATH.resolve()


def load_admissions(path: str | Path | None = None) -> dict[str, AdmissionRecord]:
    admissions_path = resolve_admissions_path(path)
    try:
        raw_records = json.loads(admissions_path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise AdmissionDataError("录取名单文件不存在") from exc
    except (OSError, json.JSONDecodeError) as exc:
        raise AdmissionDataError("录取名单文件无法读取或不是有效 JSON") from exc

    if not isinstance(raw_records, list):
        raise AdmissionDataError("录取名单必须是 JSON 数组")

    records: dict[str, AdmissionRecord] = {}
    try:
        for raw_record in raw_records:
            record = AdmissionRecord.model_validate(raw_record)
            if record.student_id in records:
                raise AdmissionDataError("录取名单包含重复学号")
            records[record.student_id] = record
    except ValidationError as exc:
        problems = []
        for error in exc.errors(include_input=False, include_url=False):
            location = ".".join(str(part) for part in error["loc"])
            problems.append(f"{location}: {error['msg']}")
        raise AdmissionDataError(
            f"录取名单校验失败：{'; '.join(problems)}"
        ) from None
    return records


def _excel_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _masked_preview(record: AdmissionRecord) -> dict:
    name = record.name[0] + "*" * max(1, len(record.name) - 1)
    return {
        "name": name,
        "studentId": "********" + record.student_id[-4:],
        "phone": "*******" + record.phone[-4:],
        "department": record.department,
        "status": record.status,
    }


def parse_admission_workbook(content: bytes) -> dict:
    if not content:
        raise AdmissionDataError("上传的 Excel 文件为空")
    if len(content) > MAX_WORKBOOK_BYTES:
        raise AdmissionDataError("Excel 文件不能超过 2 MiB")

    stream = io.BytesIO(content)
    if not zipfile.is_zipfile(stream):
        raise AdmissionDataError("上传文件不是有效的 .xlsx 工作簿")
    stream.seek(0)
    with zipfile.ZipFile(stream) as archive:
        entries = archive.infolist()
        if (
            len(entries) > MAX_WORKBOOK_ENTRIES
            or sum(entry.file_size for entry in entries) > MAX_WORKBOOK_EXPANDED_BYTES
        ):
            raise AdmissionDataError("Excel 文件展开后过大或结构异常")
        unsafe_parts = [
            entry.filename
            for entry in entries
            if "vbaproject" in entry.filename.lower()
            or entry.filename.startswith("xl/externalLinks/")
        ]
        if unsafe_parts:
            raise AdmissionDataError("Excel 文件不得包含宏或外部链接")

    stream.seek(0)
    try:
        workbook = load_workbook(stream, read_only=True, data_only=False)
    except Exception as exc:
        raise AdmissionDataError("Excel 文件无法解析") from exc

    try:
        if WORKSHEET_NAME not in workbook.sheetnames:
            raise AdmissionDataError(f"Excel 必须包含“{WORKSHEET_NAME}”工作表")
        worksheet = workbook[WORKSHEET_NAME]
        rows = worksheet.iter_rows()
        header_row = next(rows, None)
        if header_row is None:
            raise AdmissionDataError("录取名单工作表为空")
        headers = tuple(_excel_text(cell.value) for cell in header_row)
        if (
            headers[: len(WORKBOOK_COLUMNS)] != WORKBOOK_COLUMNS
            or any(headers[len(WORKBOOK_COLUMNS) :])
        ):
            raise AdmissionDataError(
                "Excel 只能包含这些表头并依次排列：" + "、".join(WORKBOOK_COLUMNS)
            )

        records: list[AdmissionRecord] = []
        errors: list[dict] = []
        seen_student_ids: set[str] = set()
        for row_number, cells in enumerate(rows, start=2):
            selected_cells = cells[: len(WORKBOOK_COLUMNS)]
            if all(cell.value is None or _excel_text(cell.value) == "" for cell in selected_cells):
                continue
            if len(records) + len(errors) >= MAX_ADMISSION_ROWS:
                raise AdmissionDataError(f"录取名单最多允许 {MAX_ADMISSION_ROWS} 行")
            if any(cell.data_type == "f" for cell in selected_cells):
                errors.append({"row": row_number, "message": "不得使用公式"})
                continue

            values = [_excel_text(cell.value) for cell in selected_cells]
            raw_record = {
                "name": values[0],
                "studentId": values[1],
                "phone": values[2],
                "department": values[3],
                "status": values[4],
            }
            try:
                record = AdmissionRecord.model_validate(raw_record)
            except ValidationError:
                errors.append({"row": row_number, "message": "字段缺失或格式不正确"})
                continue
            if record.student_id in seen_student_ids:
                errors.append({"row": row_number, "message": "学号重复"})
                continue
            seen_student_ids.add(record.student_id)
            records.append(record)

        if not records and not errors:
            raise AdmissionDataError("录取名单没有数据行")
        return {"records": records, "errors": errors}
    finally:
        workbook.close()


def create_admission_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = WORKSHEET_NAME
    worksheet.append(WORKBOOK_COLUMNS)
    worksheet.append(("张三", "202600000001", "13800000000", "嵌入式部门", "已录取"))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:E2"
    header_fill = PatternFill("solid", fgColor="DCEEFF")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column, width in zip("ABCDE", (16, 18, 18, 24, 14), strict=True):
        worksheet.column_dimensions[column].width = width
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _cleanup_previews(now: datetime) -> None:
    expired = [
        preview_id
        for preview_id, preview in _previews.items()
        if now - preview["created_at"] > PREVIEW_TTL
    ]
    for preview_id in expired:
        _previews.pop(preview_id, None)


def _store_preview(records: list[AdmissionRecord], officer: RecruitmentOfficerInfo) -> str:
    preview_id = uuid.uuid4().hex
    now = datetime.now(timezone.utc)
    with _preview_lock:
        _cleanup_previews(now)
        _previews[preview_id] = {
            "created_at": now,
            "officer": officer.username,
            "records": records,
        }
    return preview_id


def _take_preview(preview_id: str, officer: RecruitmentOfficerInfo) -> list[AdmissionRecord]:
    now = datetime.now(timezone.utc)
    with _preview_lock:
        _cleanup_previews(now)
        preview = _previews.get(preview_id)
        if not preview or preview["officer"] != officer.username:
            raise AdmissionDataError("预览已失效，请重新上传并校验")
        return list(preview["records"])


def publish_admissions(
    records: list[AdmissionRecord],
    path: str | Path | None = None,
) -> tuple[Path, Path | None]:
    global _admissions_by_student_id, _admissions_path
    admissions_path = resolve_admissions_path(path)
    payload = json.dumps(
        [record.model_dump(mode="json", by_alias=True) for record in records],
        ensure_ascii=False,
        indent=2,
    ) + "\n"

    with _admissions_lock:
        try:
            admissions_path.parent.mkdir(parents=True, exist_ok=True)
            backup_path = None
            if admissions_path.exists():
                timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
                backup_path = admissions_path.with_name(
                    f"{admissions_path.name}.previous.{timestamp}"
                )
                shutil.copy2(admissions_path, backup_path)
                os.chmod(backup_path, 0o600)

            descriptor, temporary_name = tempfile.mkstemp(
                prefix=f".{admissions_path.name}.",
                suffix=".tmp",
                dir=admissions_path.parent,
            )
            temporary_path = Path(temporary_name)
            try:
                with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as file:
                    file.write(payload)
                    file.flush()
                    os.fsync(file.fileno())
                os.chmod(temporary_path, 0o600)
                os.replace(temporary_path, admissions_path)
            finally:
                temporary_path.unlink(missing_ok=True)
        except OSError as exc:
            raise AdmissionDataError("录取名单发布失败，原名单未被替换") from exc

        _admissions_path = admissions_path
        _admissions_by_student_id = {record.student_id: record for record in records}
        return admissions_path, backup_path


def initialize_admissions_data(
    config: RecruitmentConfig,
    path: str | Path | None = None,
) -> dict[str, AdmissionRecord]:
    global _admissions_by_student_id, _admissions_path
    with _admissions_lock:
        _admissions_path = resolve_admissions_path(path)
        if config.admission_query.enabled:
            _admissions_by_student_id = load_admissions(_admissions_path)
        else:
            _admissions_by_student_id = {}
        return _admissions_by_student_id


def reset_admissions_data() -> None:
    global _admissions_by_student_id, _admissions_path
    with _admissions_lock:
        _admissions_by_student_id = None
        _admissions_path = None
    with _preview_lock:
        _previews.clear()


def get_admissions(config: RecruitmentConfig) -> dict[str, AdmissionRecord]:
    with _admissions_lock:
        if _admissions_by_student_id is None:
            return initialize_admissions_data(config)
        return _admissions_by_student_id


@router.post("/query", response_model=AdmissionQueryResponse)
def query_admission(
    query: AdmissionQuery,
    request: Request,
    config: RecruitmentConfig = Depends(get_recruitment_config),
):
    if not config.admission_query.enabled:
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            config.admission_query.notice,
        )

    client_key = get_client_ip(request)
    if not admission_query_limiter.is_allowed(client_key):
        raise HTTPException(
            status.HTTP_429_TOO_MANY_REQUESTS,
            "查询过于频繁，请稍后再试",
        )

    admission = get_admissions(config).get(query.student_id)
    if not admission or admission.phone != query.phone:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            "未找到匹配的录取结果，请检查输入或联系协会负责人",
        )

    return AdmissionQueryResponse(
        name=admission.name,
        department=admission.department,
        status=admission.status,
    )


@router.get("/manage/status")
def managed_admissions_status(
    config: RecruitmentConfig = Depends(get_recruitment_config),
    officer: RecruitmentOfficerInfo = Depends(get_current_recruitment_officer),
):
    admissions_path = resolve_admissions_path()
    if not admissions_path.exists():
        return {
            "published": False,
            "valid": True,
            "count": 0,
            "updatedAt": None,
            "queryEnabled": config.admission_query.enabled,
        }
    try:
        records = load_admissions(admissions_path)
    except AdmissionDataError:
        return {
            "published": True,
            "valid": False,
            "count": None,
            "updatedAt": datetime.fromtimestamp(
                admissions_path.stat().st_mtime,
                tz=timezone.utc,
            ),
            "queryEnabled": config.admission_query.enabled,
        }
    return {
        "published": True,
        "valid": True,
        "count": len(records),
        "updatedAt": datetime.fromtimestamp(
            admissions_path.stat().st_mtime,
            tz=timezone.utc,
        ),
        "queryEnabled": config.admission_query.enabled,
    }


@router.get("/manage/template.xlsx")
def download_admission_template(
    officer: RecruitmentOfficerInfo = Depends(get_current_recruitment_officer),
):
    return Response(
        content=create_admission_template(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": 'attachment; filename="admission-template.xlsx"',
        },
    )


@router.post("/manage/preview")
async def preview_admission_workbook(
    request: Request,
    officer: RecruitmentOfficerInfo = Depends(get_current_recruitment_officer),
):
    content_length = request.headers.get("content-length")
    if content_length and content_length.isdigit() and int(content_length) > MAX_WORKBOOK_BYTES:
        raise HTTPException(status.HTTP_413_CONTENT_TOO_LARGE, "Excel 文件不能超过 2 MiB")
    content = await request.body()
    try:
        result = parse_admission_workbook(content)
    except AdmissionDataError as exc:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_CONTENT, str(exc)) from exc

    records = result["records"]
    errors = result["errors"]
    accepted_count = sum(record.status == "已录取" for record in records)
    rejected_count = sum(record.status == "未录取" for record in records)
    if errors:
        return {
            "valid": False,
            "previewId": None,
            "summary": {
                "total": len(records) + len(errors),
                "accepted": accepted_count,
                "rejected": rejected_count,
                "errors": len(errors),
            },
            "errors": errors[:50],
            "preview": [_masked_preview(record) for record in records[:10]],
        }

    preview_id = _store_preview(records, officer)
    logger.info(
        "admissions_preview_created officer=%s count=%s",
        officer.username,
        len(records),
    )
    return {
        "valid": True,
        "previewId": preview_id,
        "summary": {
            "total": len(records),
            "accepted": accepted_count,
            "rejected": rejected_count,
            "errors": 0,
        },
        "errors": [],
        "preview": [_masked_preview(record) for record in records[:10]],
    }


@router.post("/manage/publish")
def publish_previewed_admissions(
    data: PublishAdmissionsRequest,
    config: RecruitmentConfig = Depends(get_recruitment_config),
    officer: RecruitmentOfficerInfo = Depends(get_current_recruitment_officer),
):
    if config.admission_query.enabled:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            "请先关闭录取查询，再发布新的录取名单",
        )
    try:
        records = _take_preview(data.preview_id, officer)
        _, backup_path = publish_admissions(records)
    except AdmissionDataError as exc:
        raise HTTPException(status.HTTP_409_CONFLICT, str(exc)) from exc

    with _preview_lock:
        _previews.pop(data.preview_id, None)
    logger.info(
        "admissions_published officer=%s count=%s backup=%s",
        officer.username,
        len(records),
        bool(backup_path),
    )
    return {
        "message": "录取名单已发布，录取查询仍保持关闭",
        "count": len(records),
        "backupCreated": bool(backup_path),
    }
