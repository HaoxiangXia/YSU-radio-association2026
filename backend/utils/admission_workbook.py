import io
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pydantic import ValidationError

from models.admission_list import AdmissionDataError, AdmissionRecord


MAX_WORKBOOK_BYTES = 2 * 1024 * 1024
MAX_WORKBOOK_EXPANDED_BYTES = 20 * 1024 * 1024
MAX_WORKBOOK_ENTRIES = 300
MAX_ADMISSION_ROWS = 2000
WORKSHEET_NAME = "录取名单"
WORKBOOK_COLUMNS = ("姓名", "学号", "申请手机号", "录取部门", "录取状态")


def _excel_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def masked_preview(record: AdmissionRecord) -> dict:
    name = record.name[0] + "*" * max(1, len(record.name) - 1)
    return {
        "name": name,
        "studentId": "********" + record.student_id[-4:],
        "phone": "*******" + record.phone[-4:],
        "department": record.department,
        "status": record.status,
    }


def parse_admission_workbook(content: bytes) -> dict:
    if not content:
        raise AdmissionDataError("上传的 Excel 文件为空")
    if len(content) > MAX_WORKBOOK_BYTES:
        raise AdmissionDataError("Excel 文件不能超过 2 MiB")

    stream = io.BytesIO(content)
    if not zipfile.is_zipfile(stream):
        raise AdmissionDataError("上传文件不是有效的 .xlsx 工作簿")
    stream.seek(0)
    with zipfile.ZipFile(stream) as archive:
        entries = archive.infolist()
        if (
            len(entries) > MAX_WORKBOOK_ENTRIES
            or sum(entry.file_size for entry in entries) > MAX_WORKBOOK_EXPANDED_BYTES
        ):
            raise AdmissionDataError("Excel 文件展开后过大或结构异常")
        unsafe_parts = [
            entry.filename
            for entry in entries
            if "vbaproject" in entry.filename.lower()
            or entry.filename.startswith("xl/externalLinks/")
        ]
        if unsafe_parts:
            raise AdmissionDataError("Excel 文件不得包含宏或外部链接")

    stream.seek(0)
    try:
        workbook = load_workbook(stream, read_only=True, data_only=False)
    except Exception as exc:
        raise AdmissionDataError("Excel 文件无法解析") from exc

    try:
        if WORKSHEET_NAME not in workbook.sheetnames:
            raise AdmissionDataError(f"Excel 必须包含“{WORKSHEET_NAME}”工作表")
        worksheet = workbook[WORKSHEET_NAME]
        rows = worksheet.iter_rows()
        header_row = next(rows, None)
        if header_row is None:
            raise AdmissionDataError("录取名单工作表为空")
        headers = tuple(_excel_text(cell.value) for cell in header_row)
        if (
            headers[: len(WORKBOOK_COLUMNS)] != WORKBOOK_COLUMNS
            or any(headers[len(WORKBOOK_COLUMNS) :])
        ):
            raise AdmissionDataError(
                "Excel 只能包含这些表头并依次排列：" + "、".join(WORKBOOK_COLUMNS)
            )

        records: list[AdmissionRecord] = []
        errors: list[dict] = []
        seen_student_ids: set[str] = set()
        for row_number, cells in enumerate(rows, start=2):
            selected_cells = cells[: len(WORKBOOK_COLUMNS)]
            if all(cell.value is None or _excel_text(cell.value) == "" for cell in selected_cells):
                continue
            if len(records) + len(errors) >= MAX_ADMISSION_ROWS:
                raise AdmissionDataError(f"录取名单最多允许 {MAX_ADMISSION_ROWS} 行")
            if any(cell.data_type == "f" for cell in selected_cells):
                errors.append({"row": row_number, "message": "不得使用公式"})
                continue

            values = [_excel_text(cell.value) for cell in selected_cells]
            raw_record = {
                "name": values[0],
                "studentId": values[1],
                "phone": values[2],
                "department": values[3],
                "status": values[4],
            }
            try:
                record = AdmissionRecord.model_validate(raw_record)
            except ValidationError:
                errors.append({"row": row_number, "message": "字段缺失或格式不正确"})
                continue
            if record.student_id in seen_student_ids:
                errors.append({"row": row_number, "message": "学号重复"})
                continue
            seen_student_ids.add(record.student_id)
            records.append(record)

        if not records and not errors:
            raise AdmissionDataError("录取名单没有数据行")
        return {"records": records, "errors": errors}
    finally:
        workbook.close()


def create_admission_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = WORKSHEET_NAME
    worksheet.append(WORKBOOK_COLUMNS)
    worksheet.append(("张三", "202600000001", "13800000000", "嵌入式部门", "已录取"))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:E2"
    header_fill = PatternFill("solid", fgColor="DCEEFF")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column, width in zip("ABCDE", (16, 18, 18, 24, 14), strict=True):
        worksheet.column_dimensions[column].width = width
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
